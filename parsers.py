"""
parsers.py — Lectura y normalización de archivos Excel de marcaciones y turnos.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from typing import IO

import openpyxl


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Marca:
    rut: str
    nro_personal: str
    nombre: str
    cargo: str
    seccion: str
    fecha: date
    hora: time
    cod: str       # P10 P15 P20 P25
    desc: str
    tipo: str      # Manual / Marca Reloj


@dataclass
class TurnoDay:
    codigo: int            # código empleado en turnos
    nombre: str
    contrato: str
    seccion: str
    cargo: str
    fecha: date
    raw_turno: str         # ej. "08:00 a 16:30" | "Libre" | "VAC" | etc.
    hora_inicio: time | None = None
    hora_fin: time | None = None
    es_nocturno: bool = False  # fin < inicio => cruza medianoche


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_TIME_RE = re.compile(r"(\d{1,2}):(\d{2})\s+a\s+(\d{1,2}):(\d{2})")

SKIP_TURNOS = {"libre", "vac", "enf. común", "enf. comun", "saliente", ""}


def _parse_turno_str(raw: str) -> tuple[time | None, time | None, bool]:
    """Parsea '08:00 a 16:30' → (time, time, es_nocturno)."""
    if not raw or raw.strip().lower() in SKIP_TURNOS:
        return None, None, False
    m = _TIME_RE.search(str(raw))
    if not m:
        return None, None, False
    hi = time(int(m.group(1)), int(m.group(2)))
    hf = time(int(m.group(3)), int(m.group(4)))
    nocturno = hf < hi  # cruza medianoche
    return hi, hf, nocturno


def _to_time(val) -> time | None:
    """Convierte distintos tipos de openpyxl a time."""
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        try:
            parts = val.strip().split(":")
            return time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
        except Exception:
            return None
    return None


def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


# ---------------------------------------------------------------------------
# Public parsers
# ---------------------------------------------------------------------------

def extract_store_info(file: IO[bytes]) -> dict:
    """Extrae numero y nombre de tienda desde la primera fila del Excel de marcas."""
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        division = str(row[2] or "").strip()   # ej. 'N929'
        subdir   = str(row[3] or "").strip()   # ej. 'H - Suc. 929 La Paloma'
        if division.startswith("N") and division[1:].isdigit():
            store_number = division[1:]          # '929'
            store_name   = _parse_store_name(subdir, store_number)
            return {"store_number": store_number, "store_name": store_name}
    return {"store_number": "000", "store_name": "Tienda desconocida"}


def _parse_store_name(subdir: str, store_number: str) -> str:
    """'H - Suc. 929 La Paloma' → 'La Paloma'."""
    idx = subdir.find(store_number)
    if idx != -1:
        name = subdir[idx + len(store_number):].strip()
        return name if name else subdir
    return subdir


def parse_marcas(file: IO[bytes]) -> list[Marca]:
    """Lee el archivo 'Reporte Marcas' y retorna lista de Marca."""
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    marcas: list[Marca] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        rut = str(row[7] or "").strip()
        nro = str(row[8] or "").strip()
        nombre = str(row[9] or "").strip()
        cargo = str(row[10] or "").strip()
        seccion = str(row[12] or "").strip()
        fecha = _to_date(row[19])
        hora = _to_time(row[20])
        cod = str(row[21] or "").strip()
        desc = str(row[22] or "").strip()
        tipo = str(row[25] or "").strip()

        if not fecha or not hora or not cod:
            continue

        marcas.append(Marca(
            rut=rut, nro_personal=nro, nombre=nombre,
            cargo=cargo, seccion=seccion,
            fecha=fecha, hora=hora, cod=cod, desc=desc, tipo=tipo,
        ))
    return marcas


def parse_turnos(file: IO[bytes]) -> list[TurnoDay]:
    """Lee el archivo 'ReporteTurnoResumen' (multi-hoja) y retorna lista de TurnoDay."""
    wb = openpyxl.load_workbook(file, data_only=True)
    result: list[TurnoDay] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Buscar la fila de encabezados (contiene 'Trabajador')
        header_row = None
        date_cols: list[tuple[int, date]] = []

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row[0] and str(row[0]).strip().lower() == "trabajador":
                header_row = idx
                # Columnas 7-13 (índice 7..13) tienen fechas
                for col_i in range(7, 14):
                    val = row[col_i]
                    if val is None:
                        continue
                    # Puede ser date, datetime o string "dd/mm/yyyy"
                    d = _to_date(val)
                    if d is None and isinstance(val, str):
                        try:
                            d = datetime.strptime(val.strip(), "%d/%m/%Y").date()
                        except Exception:
                            pass
                    if d:
                        date_cols.append((col_i, d))
                break

        if header_row is None or not date_cols:
            continue

        # Leer filas de datos (desde header_row+1)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row[0]:
                continue
            nombre = str(row[0]).strip()
            codigo_raw = row[1]
            codigo = int(codigo_raw) if codigo_raw and str(codigo_raw).isdigit() else 0
            contrato = str(row[2] or "").strip()
            seccion = str(row[3] or "").strip()
            cargo = str(row[4] or "").strip()

            for col_i, fecha in date_cols:
                raw = str(row[col_i] or "").strip()
                hi, hf, nocturno = _parse_turno_str(raw)
                result.append(TurnoDay(
                    codigo=codigo, nombre=nombre, contrato=contrato,
                    seccion=seccion, cargo=cargo, fecha=fecha,
                    raw_turno=raw, hora_inicio=hi, hora_fin=hf,
                    es_nocturno=nocturno,
                ))

    return result
