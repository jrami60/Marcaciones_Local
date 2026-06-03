"""
exporters.py — Generación de reportes Excel y PDF desde resultados de análisis.
Mantiene la lógica de presentación separada del servidor.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from fpdf import FPDF

# ---------------------------------------------------------------------------
# Constantes de estilo Excel
# ---------------------------------------------------------------------------
_THIN   = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_FILL_RED  = PatternFill("solid", fgColor="FFDDDD")
_FILL_YEL  = PatternFill("solid", fgColor="FFF3CD")
_FILL_GRN  = PatternFill("solid", fgColor="D4EDDA")
_FILL_BLUE = PatternFill("solid", fgColor="CCE5FF")
_FILL_GRAY = PatternFill("solid", fgColor="E9ECEF")
_FILL_HDR  = PatternFill("solid", fgColor="0053E2")

_FONT_TITLE = Font(bold=True, size=14, color="0053E2")
_FONT_SUB   = Font(size=9, color="666666")
_FONT_HDR   = Font(bold=True, size=9, color="FFFFFF")
_FONT_BOLD  = Font(bold=True, size=10)
_FONT_NORM  = Font(size=9)


def _fill_for_estado(estado: str) -> PatternFill | None:
    return {"EXCESO": _FILL_RED, "TOLERANCIA": _FILL_YEL, "SIN_COLACION": _FILL_RED}.get(estado)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def export_to_excel(rows: list[dict], store_name: str, store_number: str) -> bytes:
    """Retorna bytes de un .xlsx con hoja Resumen + hoja Detalle."""
    wb = Workbook()
    ws_res = wb.active
    ws_res.title = "Resumen"
    _build_resumen(ws_res, rows, store_name, store_number)
    _build_detalle(wb.create_sheet("Detalle"), rows, store_name, store_number)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_resumen(ws, rows: list[dict], store_name: str, store_number: str) -> None:
    total    = len(rows)
    exceso   = sum(1 for r in rows if r["estado_colacion"] == "EXCESO")
    tol      = sum(1 for r in rows if r["estado_colacion"] == "TOLERANCIA")
    ok_col   = sum(1 for r in rows if r["estado_colacion"] == "OK")
    atraso   = sum(1 for r in rows if r["estado_entrada"] == "ATRASO")
    extra    = sum(1 for r in rows if r["estado_salida"] == "EXTRA")
    sin_marc = sum(1 for r in rows if r["estado_entrada"] == "SIN_MARCA")

    ws.merge_cells("A1:F1")
    ws["A1"].value     = f"REPORTE MARCACIONES — Local {store_number} {store_name}"
    ws["A1"].font      = _FONT_TITLE
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"].value     = f"Generado: {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font      = _FONT_SUB
    ws["A2"].alignment = _CENTER
    ws.row_dimensions[2].height = 16

    tarjetas = [
        ("TOTAL REGISTROS",   total,    _FILL_BLUE),
        ("EXCESO COLACIÓN",   exceso,   _FILL_RED),
        ("EN TOLERANCIA",     tol,      _FILL_YEL),
        ("OK COLACIÓN",       ok_col,   _FILL_GRN),
        ("CON ATRASO ENTRADA",atraso,   _FILL_YEL),
        ("HICIERON EXTRA",    extra,    _FILL_BLUE),
        ("SIN MARCA ENTRADA", sin_marc, _FILL_GRAY),
    ]
    for row_n, (label, val, fill) in enumerate(tarjetas, start=4):
        ws.merge_cells(f"B{row_n}:C{row_n}")
        ws.merge_cells(f"D{row_n}:E{row_n}")
        c_lbl = ws[f"B{row_n}"]
        c_lbl.value, c_lbl.fill  = label, fill
        c_lbl.font, c_lbl.alignment, c_lbl.border = _FONT_BOLD, _LEFT, _BORDER
        c_val = ws[f"D{row_n}"]
        c_val.value     = val
        c_val.font      = Font(bold=True, size=14)
        c_val.alignment = _CENTER
        c_val.border    = _BORDER
        ws.row_dimensions[row_n].height = 22

    for col, w in [("A", 3), ("B", 22), ("C", 22), ("D", 12), ("E", 12), ("F", 3)]:
        ws.column_dimensions[col].width = w


def _build_detalle(ws, rows: list[dict], store_name: str, store_number: str) -> None:
    headers = [
        ("RUT",          12), ("Nombre",         28), ("Cargo",    18),
        ("Sección",      18), ("Fecha",           12), ("Turno",   14),
        ("Ent. Plan.",   11), ("Sal. Plan.",      11), ("Llegada", 11),
        ("Salida",       11), ("Atraso (min)",    12), ("Extra (min)", 12),
        ("Est. Entrada", 14), ("Est. Salida",     14),
        ("P1 Inicio",    10), ("P1 Fin",          10), ("P1 Dur.", 10),
        ("P1 Estado",    12),
        ("P2 Inicio",    10), ("P2 Fin",          10), ("P2 Dur.", 10),
        ("P2 Estado",    12),
        ("Col. Estado",  16), ("Alertas",         45),
    ]

    ws.merge_cells(f"A1:{_col_letter(len(headers))}1")
    ws["A1"].value     = (
        f"Detalle — Local {store_number} {store_name} — "
        f"{date.today().strftime('%d/%m/%Y')}"
    )
    ws["A1"].font      = _FONT_TITLE
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 24

    for col_i, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_i, value=header)
        cell.font, cell.fill       = _FONT_HDR, _FILL_HDR
        cell.alignment, cell.border = _CENTER, _BORDER
        ws.column_dimensions[_col_letter(col_i)].width = width
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    for row_i, r in enumerate(rows, start=3):
        p   = r.get("pausas", [])
        p1  = p[0] if len(p) > 0 else {}
        p2  = p[1] if len(p) > 1 else {}
        fill = _fill_for_estado(r.get("estado_colacion", ""))
        vals = [
            r.get("rut"),        r.get("nombre"),     r.get("cargo"),
            r.get("seccion"),    r.get("fecha"),      r.get("turno_raw"),
            r.get("turno_inicio"), r.get("turno_fin"),
            r.get("llegada"),    r.get("salida"),
            r.get("atraso_min"), r.get("extra_min"),
            r.get("estado_entrada"), r.get("estado_salida"),
            p1.get("inicio","") if p1 else "",
            p1.get("fin","")    if p1 else "",
            p1.get("duracion","") if p1 else "",
            p1.get("estado","") if p1 else "",
            p2.get("inicio","") if p2 else "",
            p2.get("fin","")    if p2 else "",
            p2.get("duracion","") if p2 else "",
            p2.get("estado","") if p2 else "",
            r.get("estado_colacion"),
            " | ".join(r.get("alertas", [])),
        ]
        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font, cell.alignment, cell.border = _FONT_NORM, _LEFT, _BORDER
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_i].height = 16


def _col_letter(n: int) -> str:
    """Convierte número de columna a letra (1→A, 27→AA)."""
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

class _PDFReport(FPDF):
    """PDF con header/footer Walmart."""
    title_text: str = ""

    def header(self):
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(0, 83, 226)
        self.cell(0, 7, self.title_text, align="C")
        self.ln(4)
        self.set_font("Helvetica", "", 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 5, f"Generado: {date.today().strftime('%d/%m/%Y')}", align="C")
        self.ln(4)
        self.set_draw_color(0, 83, 226)
        self.set_line_width(0.4)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.ln(3)

    def footer(self):
        self.set_y(-11)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(160, 160, 160)
        self.cell(0, 5, f"Walmart Chile  |  Pag. {self.page_no()}", align="C")


def export_to_pdf(rows: list[dict], store_name: str, store_number: str) -> bytes:
    """Retorna bytes de un PDF con resumen ejecutivo y tabla de detalle."""
    pdf = _PDFReport(orientation="L", unit="mm", format="A4")
    pdf.title_text = f"Reporte Marcaciones  -  Local {store_number} {store_name}"
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    _pdf_resumen_cards(pdf, rows)
    _pdf_tabla_detalle(pdf, rows)

    return bytes(pdf.output())


def _pdf_resumen_cards(pdf: _PDFReport, rows: list[dict]) -> None:
    cards = [
        ("Total registros",  len(rows),
         sum(1 for r in rows if r["estado_colacion"] == "EXCESO"),                  # col 2
         sum(1 for r in rows if r["estado_colacion"] == "TOLERANCIA"),               # col 3
         sum(1 for r in rows if r["estado_colacion"] == "OK"),                       # col 4
         sum(1 for r in rows if r["estado_entrada"] == "ATRASO"),                    # col 5
         sum(1 for r in rows if r["estado_salida"] == "EXTRA"),),                    # col 6
    ]
    labels  = ["Total", "Exceso col.", "Tolerancia", "OK col.", "Atrasos", "Extra"]
    valores = [len(rows),
               sum(1 for r in rows if r["estado_colacion"] == "EXCESO"),
               sum(1 for r in rows if r["estado_colacion"] == "TOLERANCIA"),
               sum(1 for r in rows if r["estado_colacion"] == "OK"),
               sum(1 for r in rows if r["estado_entrada"] == "ATRASO"),
               sum(1 for r in rows if r["estado_salida"] == "EXTRA")]
    colores = [
        (204, 229, 255), (255, 221, 221), (255, 243, 205),
        (212, 237, 218), (255, 243, 205), (204, 229, 255),
    ]
    w = 47
    pdf.set_font("Helvetica", "B", 8)
    for label, color in zip(labels, colores):
        pdf.set_fill_color(*color)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(w, 6, label, border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_font("Helvetica", "B", 16)
    for val, color in zip(valores, colores):
        pdf.set_fill_color(*color)
        pdf.cell(w, 9, str(val), border=1, fill=True, align="C")
    pdf.ln(14)


def _pdf_tabla_detalle(pdf: _PDFReport, rows: list[dict]) -> None:
    cols = [
        ("Nombre",      50, "L"), ("Fecha",    18, "C"),
        ("Turno",       24, "C"), ("Llegada",  16, "C"),
        ("Salida",      16, "C"), ("Atr.",     10, "C"),
        ("Ext.",        10, "C"), ("P1 Dur.",  14, "C"),
        ("P1 Est.",     18, "C"), ("P2 Dur.",  14, "C"),
        ("P2 Est.",     18, "C"), ("Colación", 20, "C"),
    ]

    # Encabezado de tabla
    pdf.set_fill_color(0, 83, 226)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 7)
    for label, w, align in cols:
        pdf.cell(w, 6, label, border=1, fill=True, align=align)
    pdf.ln()

    pdf.set_font("Helvetica", "", 6.5)
    pdf.set_text_color(40, 40, 40)

    for i, r in enumerate(rows):
        estado = r.get("estado_colacion", "OK")
        pausas = r.get("pausas", [])
        p1 = pausas[0] if len(pausas) > 0 else {}
        p2 = pausas[1] if len(pausas) > 1 else {}

        if estado == "EXCESO":
            pdf.set_fill_color(255, 221, 221)
        elif estado == "TOLERANCIA":
            pdf.set_fill_color(255, 243, 205)
        elif i % 2 == 0:
            pdf.set_fill_color(248, 249, 250)
        else:
            pdf.set_fill_color(255, 255, 255)

        dur1 = f"{p1.get('duracion','')}m" if p1 else "—"
        dur2 = f"{p2.get('duracion','')}m" if p2 else "—"
        row_vals = [
            (r.get("nombre","")[:30],     50, "L"),
            (r.get("fecha",""),            18, "C"),
            (r.get("turno_raw","")[:12],   24, "C"),
            (r.get("llegada",""),          16, "C"),
            (r.get("salida",""),           16, "C"),
            (str(r.get("atraso_min",0)),   10, "C"),
            (str(r.get("extra_min",0)),    10, "C"),
            (dur1,                         14, "C"),
            (p1.get("estado","—") if p1 else "—", 18, "C"),
            (dur2,                         14, "C"),
            (p2.get("estado","—") if p2 else "—", 18, "C"),
            (estado,                       20, "C"),
        ]
        for val, w, align in row_vals:
            pdf.cell(w, 5, str(val), border=1, fill=True, align=align)
        pdf.ln()

    total_w = sum(c[1] for c in cols)
    pdf.ln(4)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(total_w, 5, f"Total de registros: {len(rows)}", align="R")
